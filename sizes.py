from openpyxl import load_workbook


def get_row_heights(file_path, sheet_name):
    # Загружаем рабочую книгу
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[sheet_name]

    # Получаем высоты строк
    row_heights = {}
    for row in ws.row_dimensions:
        # row_dimensions индексируется с 1
        row_heights[row] = ws.row_dimensions[row].height

    return row_heights


def get_col_widths(file_path, sheet_name):
    # Загружаем рабочую книгу
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[sheet_name]

    # Получаем ширины столбцов
    col_widths = {}
    for col in ws.column_dimensions:
        # column_dimensions индексируется буквами
        col_widths[col] = ws.column_dimensions[col].width

    return col_widths


# Пример использования
file_path = 'этикетка.xlsx'  # Замените на путь к вашему файлу
sheet_name = 'Лист1'  # Замените на имя вашего листа

row_heights = get_row_heights(file_path, sheet_name)
col_widths = get_col_widths(file_path, sheet_name)

print(row_heights)
print(col_widths)