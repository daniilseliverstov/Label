from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from datetime import datetime, timedelta
from sizes import row_heights, col_widths


class Label:

    """
        Класс для создания одной этикетки в Excel-листе.

        Атрибуты:
            ROWS_PER_LABEL (int): Количество строк, занимаемых одной этикеткой.
            ws (Worksheet): Рабочий лист openpyxl, на котором создаётся этикетка.
            start_row (int): Начальная строка для размещения этикетки.
            row_offset (int): Смещение строк относительно начала листа.

        Методы:
            create(): Создаёт этикетку, применяя размеры строк, объединения ячеек, границы,
                      вставку изображений, текст и дату.
        """

    ROWS_PER_LABEL = 17

    def __init__(self, ws, start_row):

        """
        Инициализация объекта Label.

            Args:
                ws (Worksheet): Объект рабочего листа openpyxl.
                start_row (int): Начальная строка для размещения этикетки.
        """

        self.ws = ws
        self.start_row = start_row
        self.row_offset = start_row - 1

        # Константы внутри класса
        self.merge_ranges = [
            "A1:E8", "A9:B12", "C9:E12", "A13:E16", "F1:L4", "M1:O4", "P1:R4", "S1:S16",
            "F5:I8", "J5:L8", "M5:O8", "P5:R8", "F9:I12", "J9:L12", "M9:O12", "P9:R12",
            "F13:G14", "H13:I14", "J13:K14", "F15:G16", "H15:I16", "J15:K16", "L13:M16",
            "N13:N16", "O13:O16", "P13:R16"
        ]

        self.images_info = [
            ("images/Logo.png", "A2", 323.62, 108.4615384615385),
            ("images/EAC.png", "A9", 77.214, 61.53856),
            ("images/Contacts.png", "C9", 193.035, 65.38455),
        ]

        self.text_cells = [
            ("A13", "ГОСТ 16371-2014", Font(name="Times New Roman", size=16, bold=True)),
            ("F5", "Наименование упаковки", Font(name="Times New Roman", size=16, bold=True)),
            ("J5", "Цвет", Font(name="Times New Roman", size=20, bold=True)),
            ("M5", "ЗАКАЗЧИК", Font(name="Times New Roman", size=20, bold=True)),
            ("P1", "ВСЕГО УПАКОВОК", Font(name="Times New Roman", size=14, bold=True)),
            ("P9", "№ УПАКОВКИ", Font(name="Times New Roman", size=14, bold=True)),
            ("F13", "ВЫСОТА", Font(name="Times New Roman", size=14, bold=True)),
            ("H13", "ШИРИНА", Font(name="Times New Roman", size=14, bold=True)),
            ("J13", "ГЛУБИНА", Font(name="Times New Roman", size=14, bold=True)),
            ("L13", "ВЕС", Font(name="Times New Roman", size=14, bold=True)),
            ("O13", "КГ", Font(name="Times New Roman", size=14, bold=True)),
        ]

        self.center_alignment = Alignment(horizontal="center", vertical="center")

        self.thick_border = Border(
            left=Side(style="thick"),
            right=Side(style="thick"),
            top=Side(style="thick"),
            bottom=Side(style="thick")
        )

    def _apply_row_heights(self):

        """
        Применяет заданные высоты строк к диапазону строк этикетки с учётом смещения.
        """

        for r, h in row_heights.items():
            self.ws.row_dimensions[r + self.row_offset].height = h

    def _apply_merge_and_borders(self):

        """
        Объединяет ячейки в указанных диапазонах и применяет толстые границы ко всем ячейкам этих диапазонов.
        Смещает диапазоны по строкам в соответствии с позицией этикетки.
        """

        for merge_range in self.merge_ranges:
            start_cell, end_cell = merge_range.split(':')
            start_col_letter, start_row = coordinate_from_string(start_cell)
            end_col_letter, end_row = coordinate_from_string(end_cell)
            start_col = column_index_from_string(start_col_letter)
            end_col = column_index_from_string(end_col_letter)

            new_start_cell = f"{start_col_letter}{start_row + self.row_offset}"
            new_end_cell = f"{end_col_letter}{end_row + self.row_offset}"
            new_merge_range = f"{new_start_cell}:{new_end_cell}"

            self.ws.merge_cells(new_merge_range)

            for row in range(start_row + self.row_offset, end_row + self.row_offset + 1):
                for col in range(start_col, end_col + 1):
                    cell = self.ws.cell(row=row, column=col)
                    cell.border = self.thick_border

    def _insert_images(self):

        """
        Вставляет изображения в указанные ячейки с заданными размерами,
        учитывая смещение по строкам для текущей этикетки.
        """

        for path, cell, width, height in self.images_info:
            col_letter, row_num = coordinate_from_string(cell)
            new_row = row_num + self.row_offset
            new_cell = f"{col_letter}{new_row}"
            img = Image(path)
            img.width = width
            img.height = height
            self.ws.add_image(img, new_cell)

    def _set_text_cells(self):

        """
        Заполняет указанные ячейки текстом и применяет к ним заданное форматирование (шрифт и выравнивание),
        с учётом смещения по строкам.
        """

        for cell, text, font in self.text_cells:
            col_letter, row_num = coordinate_from_string(cell)
            new_row = row_num + self.row_offset
            new_cell = f"{col_letter}{new_row}"
            self.ws[new_cell] = text
            self.ws[new_cell].font = font
            self.ws[new_cell].alignment = self.center_alignment

    def _set_date(self):

        """
        Устанавливает дату (текущая дата + 7 дней) в ячейку столбца S первой строки этикетки.
        Применяет шрифт и вертикальный поворот текста.
        """

        date_cell = f"S{1 + self.row_offset}"
        label_date = (datetime.now() + timedelta(days=7)).strftime("%d.%m.%Y")
        self.ws[date_cell] = label_date
        self.ws[date_cell].font = Font(name="Times New Roman", size=26, bold=True)
        self.ws[date_cell].alignment = Alignment(
            horizontal="center", vertical="center", textRotation=90
        )

    def create(self):

        """
        Основной метод создания этикетки.
        Выполняет применение высоты строк, объединение ячеек и границ,
        вставку изображений, добавление текста и установку даты.
        """

        self._apply_row_heights()
        self._apply_merge_and_borders()
        self._insert_images()
        self._set_text_cells()
        self._set_date()


class LabelSheet:

    """
    Класс для создания Excel-файла с несколькими этикетками.

    Атрибуты:
        label_count (int): Количество этикеток для создания.
        wb (Workbook): Объект книги Excel.
        ws (Worksheet): Активный лист книги Excel.

    Методы:
        create_labels(): Создаёт все этикетки на листе.
        save(filename): Сохраняет книгу Excel в файл с указанным именем.
    """

    def __init__(self, label_count):

        """
        Инициализация объекта LabelSheet.
        Args:
            label_count (int): Количество этикеток для создания.
        """

        self.label_count = label_count
        self.wb = Workbook()
        self.ws = self.wb.active

    def _set_column_widths(self):
        """
        Устанавливает ширину столбцов согласно заданным параметрам из col_widths.
        """
        for col, width in col_widths.items():
            self.ws.column_dimensions[col].width = width

    def create_labels(self):
        """
        Создает все этикетки, размещая их друг под другом на листе.
        """
        self._set_column_widths()
        for i in range(self.label_count):
            start_row = 1 + i * Label.ROWS_PER_LABEL
            label = Label(self.ws, start_row)
            label.create()

    def save(self, filename):
        """
        Сохраняет книгу Excel в файл.

        Args:
            filename (str): Имя файла для сохранения.
        """
        self.wb.save(filename)


def main():
    """
    Основная функция программы.

    Запрашивает у пользователя количество этикеток для создания,
    создает соответствующий Excel-файл с этикетками и сохраняет его.
    Обрабатывает ошибочные вводы.
    """
    try:
        n = int(input("Введите количество этикеток: "))
        if n < 1:
            raise ValueError("Количество этикеток должно быть положительным числом.")
    except ValueError as e:
        print("Ошибка ввода:", e)
        return

    sheet = LabelSheet(n)
    sheet.create_labels()
    sheet.save("multylabel.xlsx")
    print(f"Файл multylabel.xlsx успешно сохранён с {n} этикетками.")


if __name__ == "__main__":
    main()
