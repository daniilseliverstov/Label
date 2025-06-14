from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from sizes import row_heights, col_widths
from datetime import datetime, timedelta

def set_cell(ws, cell, value, font=None, alignment=None):
    ws[cell] = value
    if font:
        ws[cell].font = font
    if alignment:
        ws[cell].alignment = alignment

def apply_border(ws, merge_ranges, border):
    for merge_range in merge_ranges:
        ws.merge_cells(merge_range)
        start_cell, end_cell = merge_range.split(':')
        start_col_letter, start_row = coordinate_from_string(start_cell)
        end_col_letter, end_row = coordinate_from_string(end_cell)
        start_col = column_index_from_string(start_col_letter)
        end_col = column_index_from_string(end_col_letter)

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border

# Создаем рабочий лист
wb = Workbook()
ws = wb.active

# Установка размеров строк и столбцов
for row, height in row_heights.items():
    ws.row_dimensions[row].height = height
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

merge_ranges = [
    "A1:E8", "A9:B12", "C9:E12", "A13:E16", "F1:L4", "M1:O4", "P1:R4", "S1:S16",
    "F5:I8", "J5:L8", "M5:O8", "P5:R8", "F9:I12", "J9:L12", "M9:O12", "P9:R12",
    "F13:G14", "H13:I14", "J13:K14", "F15:G16", "H15:I16", "J15:K16", "L13:M16",
    "N13:N16", "O13:O16", "P13:R16"
]

thick_border = Border(
    left=Side(style="thick"),
    right=Side(style="thick"),
    top=Side(style="thick"),
    bottom=Side(style="thick")
)

apply_border(ws, merge_ranges, thick_border)

# Загрузка и вставка изображений
images_info = [
    ("images/Logo.png", "A2", 323.62, 108.4615384615385),
    ("images/EAC.png", "A9", 77.214, 61.53856),
    ("images/Contacts.png", "C9", 193.035, 65.38455),
]

for path, cell, width, height in images_info:
    img = Image(path)
    img.width = width
    img.height = height
    ws.add_image(img, cell)

# Шрифты и выравнивание
center_alignment = Alignment(horizontal="center", vertical="center")

# Заполнение ячеек с текстом
text_cells = [
    ("A13", "ГОСТ 16371-2014", Font(name="Times New Roman", size=16, bold=True)),
    ("F5", "Наименование упаковки", Font(name="Times New Roman", size=16, bold=True)),
    ("J5", "Цвет", Font(name="Times New Roman", size=20, bold=True)),
    ("M5", "ЗАКАЗЧИК", Font(name="Times New Roman", size=20, bold=True)),
    ("P1", "ВСЕГО УПАКОВОК", Font(name="Times New Roman", size=14, bold=True)),
    ("P9", "№ УПАКОВКИ", Font(name="Times New Roman", size=14, bold=True)),
    ("F13", "ВЫСОТА", Font(name="Times New Roman", size=14, bold=True)),
    ("H13", "ШИРИНА", Font(name="Times New Roman", size=14, bold=True)),
    ("J13", "ГЛУБИНА", Font(name="Times New Roman", size=14, bold=True)),
    ("L13", "ВЕС", Font(name="Times New Roman", size=14, bold=True)),
    ("O13", "КГ", Font(name="Times New Roman", size=14, bold=True)),
]

for cell, text, font in text_cells:
    set_cell(ws, cell, text, font=font, alignment=center_alignment)

# Дата с поворотом текста
_date = datetime.now() + timedelta(days=7)
label_date = _date.strftime("%d.%m.%Y")
ws["S1"] = label_date
ws["S1"].font = Font(name="Times New Roman", size=26, bold=True)
ws["S1"].alignment = Alignment(horizontal="center", vertical="center", textRotation=90)

# Сохраняем файл
wb.save("label.xlsx")
