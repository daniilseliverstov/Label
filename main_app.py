import sys
import os
import re
from abc import ABC, abstractmethod
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget, QLabel,
    QLineEdit, QPushButton, QComboBox, QSpinBox, QTextEdit, QFileDialog,
    QMessageBox, QListWidget, QListWidgetItem, QInputDialog, QDialog,
    QFormLayout, QDialogButtonBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont


# Классы из order_search.py
class DataLoader(ABC):
    @abstractmethod
    def load_data(self, filename):
        pass


class ExcelDataLoader(DataLoader):
    def __init__(self):
        self.filename = None

    def load_data(self, filename=None):
        try:
            file_to_load = filename or self.filename
            if not file_to_load:
                raise ValueError("Не указан файл для загрузки")
            return pd.read_excel(file_to_load)
        except FileNotFoundError:
            raise ValueError(f"Файл '{file_to_load}' не найден.")
        except Exception as e:
            raise RuntimeError(f"Ошибка при загрузке данных: {e}")


class OrderProcessor:
    def __init__(self, data_loader: DataLoader):
        self.data_loader = data_loader

    def process_order(self, order_number):
        df = self.data_loader.load_data()
        filtered_rows = df[df['№ Заказа'].astype(str) == str(order_number)]

        if filtered_rows.empty:
            return f"Заказ №{order_number} не найден."

        first_row = filtered_rows.iloc[0]
        info_extractor = InfoExtractor(first_row)
        extracted_info = info_extractor.extract()
        return extracted_info


class InfoExtractor:
    def __init__(self, row):
        self.row = row

    def extract(self):
        info = {
            'store_application_number': self._extract_store_application(),
            'client': self._extract_client(),
            'full_name': self._extract_full_name(),
            'item_name': self._extract_item_name(),
            'dimensions': self._extract_dimensions(),
            'carcase': self._extract_carcase(),
            'extra_component': self._extract_extra_component(),
            'facade': self._extract_facade(),
            'weight': self._extract_weight(),
        }
        return OrderInfo(**info)

    def _extract_store_application(self):
        return self.row.get('№ магазина / заявка', '')

    def _extract_client(self):
        return self.row.get('Клиент', '')

    def _extract_full_name(self):
        return self.row.get('Наименование', '')

    def _extract_item_name(self):
        match = re.match(r'(.*?)(\d+)[xхХХ*×]', self.row.get('Наименование', ''))
        return match.group(1).strip() if match else ''

    def _extract_dimensions(self):
        dimensions_match = re.search(r'(\d+)\s*[xхХХ*×]\s*(\d+)\s*[xхХХ*×]\s*(\d+)', self.row.get('Наименование', ''))
        return tuple(map(int, dimensions_match.groups())) if dimensions_match else ()

    def _extract_carcase(self):
        raw_carcase = self.row.get('Корпус', '').split('/')
        words = {re.match(r'\D+', p.strip()).group().strip() for p in raw_carcase if p.strip()}
        return '/'.join(words)

    def _extract_extra_component(self):
        component = self.row.get('Профиль /            Доп. Элементы', '')
        return None if component in ['-', ''] else component

    def _extract_facade(self):
        facade = self.row.get('Фасад', '')
        return None if facade in ['-', ''] else facade

    def _extract_weight(self):
        weight = self.row.get('ВЕС, КГ', '')
        return float(weight) if isinstance(weight, (float, int)) else None


class OrderInfo:
    def __init__(self, **kwargs):
        self.store_application_number = kwargs.get('store_application_number', '')
        self.client = kwargs.get('client', '')
        self.full_name = kwargs.get('full_name', '')
        self.item_name = kwargs.get('item_name', '')
        self.dimensions = kwargs.get('dimensions', ())
        self.carcase = kwargs.get('carcase', '')
        self.extra_component = kwargs.get('extra_component', None)
        self.facade = kwargs.get('facade', None)
        self.weight = kwargs.get('weight', None)

    def format_output(self):
        output = [
            f"✅ Номер заказа: {self.store_application_number}",
            f"✅ Магазин / Заявка: {self.client}",
            f"✅ Полное наименование: {self.full_name}",
            f"✅ Наименование изделия: {self.item_name}"
        ]
        if len(self.dimensions) >= 3:
            output.extend([
                f"✅ Ширина: {self.dimensions[0]} мм",
                f"✅ Высота: {self.dimensions[1]} мм",
                f"✅ Глубина: {self.dimensions[2]} мм"
            ])
        output.append(f"✅ Корпус: {self.carcase}")
        output.append(f"✅ Дополнительный компонент: {self.extra_component or 'нет данных'}")
        output.append(f"✅ Фасад: {self.facade or 'нет данных'}")
        if self.weight is not None:
            output.append(f"✅ Вес: {int(self.weight)} кг")

        return "\n".join(output)


# Константы для размеров ячеек
from sizes import row_heights, col_widths


class LabelEditorDialog(QDialog):
    def __init__(self, label_data, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Редактирование данных этикетки")
        self.label_data = label_data
        self.init_ui()

    def init_ui(self):
        layout = QFormLayout(self)

        # Поля для редактирования
        self.item_name_edit = QLineEdit(self.label_data['item_name'])
        self.width_edit = QLineEdit(
            str(self.label_data['dimensions'][0]) if len(self.label_data['dimensions']) > 0 else "0")
        self.height_edit = QLineEdit(str(self.label_data['dimensions'][1])) if len(
            self.label_data['dimensions']) > 1 else "0"
        self.depth_edit = QLineEdit(str(self.label_data['dimensions'][2])) if len(
            self.label_data['dimensions']) > 2 else "0"
        self.weight_edit = QLineEdit(str(int(self.label_data['weight'])) if self.label_data['weight'] else "0")
        self.store_number_edit = QLineEdit(self.label_data['store_number'])
        self.client_edit = QLineEdit(self.label_data['client'])
        self.order_number_edit = QLineEdit(self.label_data['order_number'])

        # Определяем значение для компонента
        label_type = self.label_data['label_type'].upper()
        if label_type == "КОРПУС":
            component_value = self.label_data['carcase']
        elif label_type == "ОРГАЛИТ":
            component_value = "БЕЛЫЙ"
        elif label_type in ["ФАСАДЫ МДФ", "ФАСАДЫ ПЛАСТИК"]:
            component_value = self.label_data.get('facade', '')
        else:
            component_value = self.label_data.get('extra_component', '')

        self.component_edit = QLineEdit(str(component_value) if component_value is not None else "")

        # Добавляем поля в форму
        layout.addRow("Наименование изделия:", self.item_name_edit)
        layout.addRow("Ширина (мм):", self.width_edit)
        layout.addRow("Высота (мм):", self.height_edit)
        layout.addRow("Глубина (мм):", self.depth_edit)
        layout.addRow("Вес (кг):", self.weight_edit)
        layout.addRow("Номер заказа:", self.order_number_edit)
        layout.addRow("Номер магазина:", self.store_number_edit)
        layout.addRow("Клиент:", self.client_edit)
        layout.addRow("Компонент:", self.component_edit)

        # Кнопки
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

        layout.addRow(buttons)

    def get_edited_data(self):
        # Обновляем данные на основе введенных значений
        self.label_data['item_name'] = self.item_name_edit.text()

        try:
            width = int(self.width_edit.text())
            height = int(self.height_edit.text())
            depth = int(self.depth_edit.text())
            self.label_data['dimensions'] = (width, height, depth)
        except ValueError:
            pass

        try:
            self.label_data['weight'] = float(self.weight_edit.text())
        except ValueError:
            self.label_data['weight'] = None

        self.label_data['store_number'] = self.store_number_edit.text()
        self.label_data['client'] = self.client_edit.text()
        self.label_data['order_number'] = self.order_number_edit.text()

        # Обновляем компонент в зависимости от типа этикетки
        label_type = self.label_data['label_type'].upper()
        if label_type == "КОРПУС":
            self.label_data['carcase'] = self.component_edit.text()
        elif label_type == "ОРГАЛИТ":
            pass  # Оставляем "БЕЛЫЙ"
        elif label_type in ["ФАСАДЫ МДФ", "ФАСАДЫ ПЛАСТИК"]:
            self.label_data['facade'] = self.component_edit.text()
        else:
            self.label_data['extra_component'] = self.component_edit.text()

        return self.label_data


class Label:
    ROWS_PER_LABEL = 17

    def __init__(self, ws, start_row, label_data):
        self.ws = ws
        self.start_row = start_row
        self.row_offset = start_row - 1
        self.label_data = label_data

        self.merge_ranges = [
            "A1:E8", "A9:B12", "C9:E12", "A13:E16", "F1:L4", "M1:O4", "P1:R4", "S1:S16",
            "F5:I8", "J5:L8", "M5:O8", "P5:R8", "F9:I12", "J9:L12", "M9:O12", "P9:R12",
            "F13:G14", "H13:I14", "J13:K14", "F15:G16", "H15:I16", "J15:K16", "L13:M16",
            "N13:N16", "O13:O16", "P13:R16"
        ]

        self.images_info = [
            ("images/Logo.png", "A2", 323.62, 108.4615384615385),
            ("images/EAC.png", "A9", 77.214, 61.53856),
            ("images/Contacts.png", "C9", 193.035, 65.38455),
        ]

        self.center_alignment = Alignment(horizontal="center", vertical="center")

        self.thick_border = Border(
            left=Side(style="thick"),
            right=Side(style="thick"),
            top=Side(style="thick"),
            bottom=Side(style="thick")
        )

    def _apply_row_heights(self):
        for r, h in row_heights.items():
            self.ws.row_dimensions[r + self.row_offset].height = h

    def _apply_merge_and_borders(self):
        for merge_range in self.merge_ranges:
            start_cell, end_cell = merge_range.split(':')
            start_col_letter, start_row = coordinate_from_string(start_cell)
            end_col_letter, end_row = coordinate_from_string(end_cell)
            start_col = column_index_from_string(start_col_letter)
            end_col = column_index_from_string(end_col_letter)

            new_start_cell = f"{start_col_letter}{start_row + self.row_offset}"
            new_end_cell = f"{end_col_letter}{end_row + self.row_offset}"
            new_merge_range = f"{new_start_cell}:{new_end_cell}"

            self.ws.merge_cells(new_merge_range)

            for row in range(start_row + self.row_offset, end_row + self.row_offset + 1):
                for col in range(start_col, end_col + 1):
                    cell = self.ws.cell(row=row, column=col)
                    cell.border = self.thick_border

    def _insert_images(self):
        for path, cell, width, height in self.images_info:
            if not os.path.exists(path):
                continue

            try:
                col_letter, row_num = coordinate_from_string(cell)
                new_row = row_num + self.row_offset
                new_cell = f"{col_letter}{new_row}"
                img = Image(path)
                img.width = width
                img.height = height
                self.ws.add_image(img, new_cell)
            except Exception as e:
                print(f"Ошибка при вставке изображения {path}: {e}")

    def _set_text_cells(self):
        data = self.label_data
        dimensions = data.get('dimensions', (0, 0, 0))

        # Определяем значение для J9
        label_type = data['label_type'].upper()
        if label_type == "КОРПУС":
            j9_value = data['carcase']
        elif label_type == "ОРГАЛИТ":
            j9_value = "БЕЛЫЙ"
        elif label_type in ["ФАСАДЫ МДФ", "ФАСАДЫ ПЛАСТИК"]:
            j9_value = data.get('facade', '')
        else:
            j9_value = data.get('extra_component', '')

        # Формируем текст для M9
        m9_text = f"{data.get('client', '')}/{data.get('store_number', '')}"

        text_cells = [
            ("A13", "ГОСТ 16371-2014", 16),
            ("F1", data.get('item_name', ''), 16),
            ("F9", data['label_type'].upper(), 24),
            ("J9", j9_value, 16),
            ("F15", str(dimensions[1]) if len(dimensions) > 1 else "", 14),
            ("H15", str(dimensions[0]) if len(dimensions) > 0 else "", 14),
            ("J15", str(dimensions[2]) if len(dimensions) > 2 else "", 14),
            ("N13", str(int(data['weight'])) if data.get('weight') else "", 14),
            ("M1", f"№ {data.get('order_number', '')}", 20),
            ("M9", m9_text, 14),
            ("P5", str(data.get('package_total', 1)), 20),
            ("P13", str(data.get('package_num', 1)), 20),
            ("F5", "Наименование упаковки", 16),
            ("J5", "Цвет", 20),
            ("M5", "ЗАКАЗЧИК", 20),
            ("P1", "ВСЕГО УПАКОВОК", 14),
            ("P9", "№ УПАКОВКИ", 14),
            ("F13", "ВЫСОТА", 14),
            ("H13", "ШИРИНА", 14),
            ("J13", "ГЛУБИНА", 14),
            ("L13", "ВЕС", 14),
            ("O13", "КГ", 14),
        ]

        for cell, text, size in text_cells:
            if not text:
                continue

            col_letter, row_num = coordinate_from_string(cell)
            new_row = row_num + self.row_offset
            new_cell = f"{col_letter}{new_row}"

            self.ws[new_cell] = text
            self.ws[new_cell].font = Font(name="Times New Roman", size=size, bold=True)
            self.ws[new_cell].alignment = self.center_alignment

    def _set_date(self):
        date_cell = f"S{1 + self.row_offset}"
        label_date = (datetime.now() + timedelta(days=7)).strftime("%d.%m.%Y")
        self.ws[date_cell] = label_date
        self.ws[date_cell].font = Font(name="Times New Roman", size=26, bold=True)
        self.ws[date_cell].alignment = Alignment(
            horizontal="center", vertical="center", textRotation=90
        )

    def create(self):
        self._apply_row_heights()
        self._apply_merge_and_borders()
        self._insert_images()
        self._set_text_cells()
        self._set_date()


class LabelSheet:
    def __init__(self, labels_data):
        self.labels_data = labels_data
        self.wb = Workbook()
        self.ws = self.wb.active

    def _set_column_widths(self):
        for col, width in col_widths.items():
            self.ws.column_dimensions[col].width = width

    def create_labels(self):
        self._set_column_widths()
        package_num = 1

        for label_info in self.labels_data['labels']:
            label_type = label_info['label_type']
            count = label_info['count']

            for _ in range(count):
                label_data = {
                    'label_type': label_type,
                    'item_name': label_info['item_name'],
                    'dimensions': label_info['dimensions'],
                    'weight': label_info['weight'],
                    'store_number': label_info['store_number'],
                    'client': label_info['client'],
                    'carcase': label_info['carcase'],
                    'extra_component': label_info['extra_component'],
                    'facade': label_info.get('facade', ''),
                    'order_number': label_info['order_number'],
                    'package_total': self.labels_data['package_total'],
                    'package_num': package_num
                }

                start_row = 1 + (package_num - 1) * Label.ROWS_PER_LABEL
                label = Label(self.ws, start_row, label_data)
                label.create()
                package_num += 1

    def save(self, filename):
        try:
            self.wb.save(filename)
            return True
        except Exception as e:
            print(f"Ошибка при сохранении файла: {e}")
            return False


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Генератор этикеток")
        self.setMinimumSize(800, 600)

        self.excel_file_path = None
        self.order_info = None
        self.label_types = ["КОРПУС", "ФАСАДЫ МДФ", "ФАСАДЫ ПЛАСТИК", "Профиль/доп элемент", "ОРГАЛИТ"]
        self.labels_to_create = []

        self.init_ui()
        self.setup_connections()

    def init_ui(self):
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        self.main_layout = QVBoxLayout()
        self.central_widget.setLayout(self.main_layout)

        # Блок загрузки файла
        self.file_group = QWidget()
        file_layout = QHBoxLayout(self.file_group)

        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("Укажите путь к файлу раскроя...")
        file_layout.addWidget(self.file_path_edit)

        self.browse_btn = QPushButton("Обзор...")
        file_layout.addWidget(self.browse_btn)

        self.main_layout.addWidget(self.file_group)

        # Блок поиска заказа
        self.search_group = QWidget()
        search_layout = QHBoxLayout(self.search_group)

        self.order_number_edit = QLineEdit()
        self.order_number_edit.setPlaceholderText("Введите номер заказа...")
        search_layout.addWidget(self.order_number_edit)

        self.search_btn = QPushButton("Найти заказ")
        search_layout.addWidget(self.search_btn)

        self.main_layout.addWidget(self.search_group)

        # Информация о заказе
        self.order_info_text = QTextEdit()
        self.order_info_text.setReadOnly(True)
        self.order_info_text.setFont(QFont("Arial", 10))
        self.main_layout.addWidget(self.order_info_text)

        # Блок добавления этикеток
        self.label_group = QWidget()
        label_layout = QHBoxLayout(self.label_group)

        self.label_type_combo = QComboBox()
        self.label_type_combo.addItems(self.label_types)
        label_layout.addWidget(self.label_type_combo)

        self.label_count_spin = QSpinBox()
        self.label_count_spin.setRange(1, 100)
        self.label_count_spin.setValue(1)
        label_layout.addWidget(QLabel("Количество:"))
        label_layout.addWidget(self.label_count_spin)

        self.add_label_btn = QPushButton("Добавить")
        label_layout.addWidget(self.add_label_btn)

        self.edit_types_btn = QPushButton("Редактировать типы")
        label_layout.addWidget(self.edit_types_btn)

        self.main_layout.addWidget(self.label_group)

        # Список добавленных этикеток
        self.labels_list = QListWidget()
        self.labels_list.itemDoubleClicked.connect(self.edit_label)
        self.main_layout.addWidget(self.labels_list)

        # Кнопки управления
        self.control_group = QWidget()
        control_layout = QHBoxLayout(self.control_group)

        self.clear_btn = QPushButton("Очистить список")
        control_layout.addWidget(self.clear_btn)

        self.create_btn = QPushButton("Создать этикетки")
        self.create_btn.setStyleSheet("background-color: #4CAF50; color: white;")
        control_layout.addWidget(self.create_btn)

        self.main_layout.addWidget(self.control_group)

    def setup_connections(self):
        self.browse_btn.clicked.connect(self.browse_file)
        self.search_btn.clicked.connect(self.search_order)
        self.add_label_btn.clicked.connect(self.add_label)
        self.edit_types_btn.clicked.connect(self.edit_label_types)
        self.clear_btn.clicked.connect(self.clear_labels)
        self.create_btn.clicked.connect(self.create_labels)

    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите файл раскроя",
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        if file_path:
            self.file_path_edit.setText(file_path)
            self.excel_file_path = file_path

    def search_order(self):
        order_number = self.order_number_edit.text().strip()
        if not order_number:
            self.show_error("Введите номер заказа")
            return

        if not self.excel_file_path or not os.path.exists(self.excel_file_path):
            self.show_error("Сначала укажите корректный файл раскроя")
            return

        try:
            loader = ExcelDataLoader()
            loader.filename = self.excel_file_path
            processor = OrderProcessor(loader)
            self.order_info = processor.process_order(order_number)

            if isinstance(self.order_info, str):
                self.show_error(self.order_info)
            else:
                self.order_info_text.setText(self.order_info.format_output())
                self.show_info("Данные заказа успешно загружены")

        except Exception as e:
            self.show_error(f"Ошибка при поиске заказа: {str(e)}")

    def add_label(self):
        if not self.order_info or isinstance(self.order_info, str):
            self.show_error("Сначала найдите корректный заказ")
            return

        label_type = self.label_type_combo.currentText()
        count = self.label_count_spin.value()

        # Создаем словарь с данными для этикетки
        label_data = {
            'label_type': label_type,
            'count': count,
            'item_name': getattr(self.order_info, 'item_name', ''),
            'dimensions': getattr(self.order_info, 'dimensions', (0, 0, 0)),
            'weight': getattr(self.order_info, 'weight', None),
            'store_number': getattr(self.order_info, 'store_application_number', ''),
            'client': getattr(self.order_info, 'client', ''),
            'carcase': getattr(self.order_info, 'carcase', ''),
            'extra_component': getattr(self.order_info, 'extra_component', ''),
            'facade': getattr(self.order_info, 'facade', ''),
            'order_number': self.order_number_edit.text().strip(),
        }

        # Открываем диалог редактирования
        dialog = LabelEditorDialog(label_data, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            edited_data = dialog.get_edited_data()
            self.labels_to_create.append(edited_data)
            self.update_labels_list()

            # Сбрасываем счетчик
            self.label_count_spin.setValue(1)

    def edit_label(self, item):
        # Получаем индекс выбранного элемента
        index = self.labels_list.row(item)
        if index < 0 or index >= len(self.labels_to_create):
            return

        # Открываем диалог редактирования
        dialog = LabelEditorDialog(self.labels_to_create[index], self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.labels_to_create[index] = dialog.get_edited_data()
            self.update_labels_list()

    def update_labels_list(self):
        self.labels_list.clear()
        for label in self.labels_to_create:
            item = QListWidgetItem(
                f"{label['label_type']} - {label['count']} шт. | "
                f"Название: {label['item_name']} | "
                f"Размеры: {label['dimensions'][0]}x{label['dimensions'][1]}x{label['dimensions'][2]}"
            )
            self.labels_list.addItem(item)

    def edit_label_types(self):
        new_type, ok = QInputDialog.getText(
            self,
            "Редактирование типов",
            "Введите новый тип этикетки:",
            QLineEdit.EchoMode.Normal,
            ""
        )

        if ok and new_type.strip():
            if new_type.strip().upper() not in [t.upper() for t in self.label_types]:
                self.label_types.append(new_type.strip())
                self.label_type_combo.addItem(new_type.strip())
                self.show_info(f"Тип '{new_type}' добавлен")
            else:
                self.show_error("Такой тип уже существует")

    def clear_labels(self):
        self.labels_to_create = []
        self.labels_list.clear()

    def create_labels(self):
        if not self.order_info or isinstance(self.order_info, str):
            self.show_error("Нет данных заказа для создания этикеток")
            return

        if not self.labels_to_create:
            self.show_error("Нет этикеток для создания")
            return

        # Подсчитываем общее количество этикеток
        total_labels = sum(label['count'] for label in self.labels_to_create)

        # Собираем данные для этикеток
        labels_data = {
            'labels': self.labels_to_create,
            'package_total': total_labels
        }

        # Запрашиваем путь для сохранения
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить файл этикеток",
            "multylabel.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            sheet = LabelSheet(labels_data)
            sheet.create_labels()

            if sheet.save(file_path):
                self.show_info(f"Файл успешно сохранён с {total_labels} этикетками")
                self.clear_labels()
            else:
                self.show_error("Не удалось сохранить файл")

        except Exception as e:
            self.show_error(f"Ошибка при создании файла: {str(e)}")

    def show_error(self, message):
        QMessageBox.critical(self, "Ошибка", message)

    def show_info(self, message):
        QMessageBox.information(self, "Информация", message)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle('Fusion')

    window = MainWindow()
    window.show()

    sys.exit(app.exec())