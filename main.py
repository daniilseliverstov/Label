import sys
import os
import json
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QLineEdit, QFileDialog, QMessageBox, QComboBox, QSpinBox,
    QTextEdit, QListWidget, QListWidgetItem, QCheckBox, QFormLayout, QGroupBox
)
from PyQt6.QtCore import Qt

import pandas as pd
import re
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# --- sizes.py content (needed for multy.py) ---
row_heights = {
    1: 3, 2: 3, 3: 3, 4: 3, 5: 3, 6: 3, 7: 3, 8: 3,
    9: 3, 10: 3, 11: 3, 12: 3, 13: 3, 14: 3, 15: 3, 16: 3
}
col_widths = {
    "A": 8, "B": 8, "C": 8, "D": 8, "E": 8,
    "F": 8, "G": 8, "H": 8, "I": 8, "J": 8,
    "K": 8, "L": 8, "M": 8, "N": 8, "O": 8,
    "P": 8, "Q": 8, "R": 8, "S": 8
}


# --- Order processing classes (adapted from order_search.py) ---

class OrderNotFoundError(Exception):
    pass

class ExcelDataLoader:
    """Класс для загрузки данных из Excel-файлов"""

    def __init__(self, filename):
        self.filename = filename

    def load_data(self):
        if not os.path.exists(self.filename):
            raise FileNotFoundError(f"Файл '{self.filename}' не найден.")
        try:
            return pd.read_excel(self.filename)
        except Exception as e:
            raise RuntimeError(f"Ошибка при загрузке данных: {e}")

class InfoExtractor:
    def __init__(self, row):
        self.row = row

    def extract(self):
        info = {
            'store_application_number': self._extract_store_application(),
            'client': self._extract_client(),
            'full_name': self._extract_full_name(),
            'item_name': self._extract_item_name(),
            'dimensions': self._extract_dimensions(),
            'carcase': self._extract_carcase(),
            'extra_component': self._extract_extra_component(),
            'facade': self._extract_facade(),
            'weight': self._extract_weight(),
        }
        return OrderInfo(**info)

    def _extract_store_application(self):
        return self.row.get('№ магазина / заявка', '')

    def _extract_client(self):
        return self.row.get('Клиент', '')

    def _extract_full_name(self):
        return self.row.get('Наименование', '')

    def _extract_item_name(self):
        name = self.row.get('Наименование', '')
        match = re.match(r'(.*?)(\d+)[xхХХ*×]', name)
        return match.group(1).strip() if match else name

    def _extract_dimensions(self):
        name = self.row.get('Наименование', '')
        dimensions_match = re.search(r'(\d+)\s*[xхХХ*×]\s*(\d+)\s*[xхХХ*×]\s*(\d+)', name)
        return tuple(map(int, dimensions_match.groups())) if dimensions_match else (0, 0, 0)

    def _extract_carcase(self):
        raw_carcase = self.row.get('Корпус', '')
        parts = [p.strip() for p in raw_carcase.split('/') if p.strip()]
        words = []
        for p in parts:
            m = re.match(r'\D+', p)
            if m:
                words.append(m.group().strip())
        return '/'.join(words)

    def _extract_extra_component(self):
        component = self.row.get('Профиль /            Доп. Элементы', '')
        return None if component in ['-', ''] else component

    def _extract_facade(self):
        facade = self.row.get('Фасад', '')
        return None if facade in ['-', ''] else facade

    def _extract_weight(self):
        weight = self.row.get('ВЕС, КГ', '')
        try:
            return float(weight)
        except (ValueError, TypeError):
            return None

class OrderInfo:
    def __init__(self, **kwargs):
        self.store_application_number = kwargs.get('store_application_number', '')
        self.client = kwargs.get('client', '')
        self.full_name = kwargs.get('full_name', '')
        self.item_name = kwargs.get('item_name', '')
        self.dimensions = kwargs.get('dimensions', (0, 0, 0))
        self.carcase = kwargs.get('carcase', '')
        self.extra_component = kwargs.get('extra_component', None)
        self.facade = kwargs.get('facade', None)
        self.weight = kwargs.get('weight', None)

class OrderProcessor:
    def __init__(self, filename):
        self.loader = ExcelDataLoader(filename)
        self.df = None

    def load(self):
        self.df = self.loader.load_data()

    def process_order(self, order_number) -> OrderInfo:
        if self.df is None:
            self.load()
        filtered_rows = self.df[self.df['№ Заказа'].astype(str) == str(order_number)]
        if filtered_rows.empty:
            raise OrderNotFoundError(f"Заказ №{order_number} не найден.")
        first_row = filtered_rows.iloc[0]
        extractor = InfoExtractor(first_row)
        return extractor.extract()


# --- Label generation (based on multy.py) ---

def set_cell(ws, cell, value, font=None, alignment=None):
    ws[cell] = value
    if font:
        ws[cell].font = font
    if alignment:
        ws[cell].alignment = alignment

def apply_border(ws, merge_ranges, border, row_offset=0):
    for merge_range in merge_ranges:
        start_cell, end_cell = merge_range.split(':')
        start_col_letter, start_row = coordinate_from_string(start_cell)
        end_col_letter, end_row = coordinate_from_string(end_cell)
        start_col = column_index_from_string(start_col_letter)
        end_col = column_index_from_string(end_col_letter)

        # Сдвигаем строки на row_offset
        new_start_cell = f"{start_col_letter}{start_row + row_offset}"
        new_end_cell = f"{end_col_letter}{end_row + row_offset}"
        new_merge_range = f"{new_start_cell}:{new_end_cell}"

        ws.merge_cells(new_merge_range)

        for row in range(start_row + row_offset, end_row + row_offset + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border

row_heights = {
    1: 15, 2: 15, 3: 15, 4: 15, 5: 15, 6: 15, 7: 15, 8: 15,
    9: 15, 10: 15, 11: 15, 12: 15, 13: 15, 14: 15, 15: 15, 16: 15, 17: 15
}
col_widths = {
    "A": 4, "B": 4, "C": 4, "D": 4, "E": 4,
    "F": 12, "G": 12, "H": 12, "I": 12, "J": 12,
    "K": 12, "L": 12, "M": 12, "N": 12, "O": 12,
    "P": 12, "Q": 12, "R": 12, "S": 12
}

def create_label(ws, start_row, info: OrderInfo, label_type: str, total_count: int, current_number: int, editable_values=None):
    """
    Создает одну этикетку, начиная со строки start_row, с параметрами из info и label_type
    editable_values - dict с возможными изменёнными значениями для ячеек
    """

    row_offset = start_row - 1

    # Размеры строк для этикетки
    for r, h in row_heights.items():
        ws.row_dimensions[r + row_offset].height = h

    # Размеры столбцов (ставим только один раз в начале, т.к. Столбцы не меняются)
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    merge_ranges = [
        "A1:E8", "A9:B12", "C9:E12", "A13:E16", "F1:L4", "M1:O4", "P1:R4", "S1:S16",
        "F5:I8", "J5:L8", "M5:O8", "P5:R8", "F9:I12", "J9:L12", "M9:O12", "P9:R12",
        "F13:G14", "H13:I14", "J13:K14", "F15:G16", "H15:I16", "J15:K16", "L13:M16",
        "N13:N16", "O13:O16", "P13:R16"
    ]

    thick_border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick")
    )

    apply_border(ws, merge_ranges, thick_border, row_offset=row_offset)

    center_alignment = Alignment(horizontal="center", vertical="center")

    # Тексты, которые не меняются
    text_cells = [
        ("A13", "ГОСТ 16371-2014", Font(name="Times New Roman", size=16, bold=True)),
        ("F5", "Наименование упаковки", Font(name="Times New Roman", size=16, bold=True)),
        ("J5", "Цвет", Font(name="Times New Roman", size=20, bold=True)),
        ("M5", "ЗАКАЗЧИК", Font(name="Times New Roman", size=20, bold=True)),
        ("P1", "ВСЕГО УПАКОВОК", Font(name="Times New Roman", size=14, bold=True)),
        ("P9", "№ УПАКОВКИ", Font(name="Times New Roman", size=14, bold=True)),
        ("F13", "ВЫСОТА", Font(name="Times New Roman", size=14, bold=True)),
        ("H13", "ШИРИНА", Font(name="Times New Roman", size=14, bold=True)),
        ("J13", "ГЛУБИНА", Font(name="Times New Roman", size=14, bold=True)),
        ("L13", "ВЕС", Font(name="Times New Roman", size=14, bold=True)),
        ("O13", "КГ", Font(name="Times New Roman", size=14, bold=True)),
    ]

    for cell, text, font in text_cells:
        col_letter, row_num = coordinate_from_string(cell)
        new_row = row_num + row_offset
        new_cell = f"{col_letter}{new_row}"
        set_cell(ws, new_cell, text, font=font, alignment=center_alignment)

    # Дата с поворотом текста
    _date = datetime.now() + timedelta(days=7)
    label_date = _date.strftime("%d.%m.%Y")
    s_cell = f"S{1 + row_offset}"
    ws[s_cell] = label_date
    ws[s_cell].font = Font(name="Times New Roman", size=26, bold=True)
    ws[s_cell].alignment = Alignment(horizontal="center", vertical="center", textRotation=90)

    # Вставляем данные из info и label_type с возможностью редактирования
    # Если editable_values переданы, берем из них, иначе из info

    def val(key, default=''):
        if editable_values and key in editable_values:
            return editable_values[key]
        return default

    # F1 - Наименование изделия
    f1 = f"F{1 + row_offset}"
    set_cell(ws, f1, val('item_name', info.item_name), font=Font(name="Times New Roman", size=20, bold=True), alignment=center_alignment)

    # F15 - Ширина
    f15 = f"F{15 + row_offset}"
    set_cell(ws, f15, val('width', info.dimensions[0]), font=Font(name="Times New Roman", size=14), alignment=center_alignment)

    # H15 - Высота
    h15 = f"H{15 + row_offset}"
    set_cell(ws, h15, val('height', info.dimensions[1]), font=Font(name="Times New Roman", size=14), alignment=center_alignment)

    # J15 - Глубина
    j15 = f"J{15 + row_offset}"
    set_cell(ws, j15, val('depth', info.dimensions[2]), font=Font(name="Times New Roman", size=14), alignment=center_alignment)

    # J9 - Корпус / Доп. компонент / БЕЛЫЙ
    j9 = f"J{9 + row_offset}"
    if label_type.upper() == 'КОРПУС':
        j9_val = val('carcase', info.carcase)
    elif label_type.upper() == 'ОРГАЛИТ':
        j9_val = 'БЕЛЫЙ'
    else:
        j9_val = val('extra_component', info.extra_component or '')
    set_cell(ws, j9, j9_val.upper(), font=Font(name="Times New Roman", size=14, bold=True), alignment=center_alignment)

    # N13 - Вес
    n13 = f"N{13 + row_offset}"
    w = val('weight', info.weight)
    w_str = str(int(w)) if w else ''
    set_cell(ws, n13, w_str, font=Font(name="Times New Roman", size=14), alignment=center_alignment)

    # M1 - № заказа
    m1 = f"M{1 + row_offset}"
    set_cell(ws, m1, f"№ {val('store_application_number', info.store_application_number)}", font=Font(name="Times New Roman", size=14, bold=True), alignment=center_alignment)

    # M9 - Клиент / № заявки
    m9 = f"M{9 + row_offset}"
    client = val('client', info.client)
    store_app = val('store_application_number', info.store_application_number)
    m9_val = f"{client}/{store_app}"
    set_cell(ws, m9, m9_val, font=Font(name="Times New Roman", size=12), alignment=center_alignment)

    # P5 - Общее количество упаковок
    p5 = f"P{5 + row_offset}"
    set_cell(ws, p5, str(total_count), font=Font(name="Times New Roman", size=14, bold=True), alignment=center_alignment)

    # R13 - Порядковый номер упаковки
    r13 = f"R{13 + row_offset}"
    set_cell(ws, r13, str(current_number), font=Font(name="Times New Roman", size=14, bold=True), alignment=center_alignment)

# --- Main PyQt6 application ---

from PyQt6.QtGui import QFont

class LabelTypeItem:
    def __init__(self, name):
        self.name = name

class LabelTypeManager(QWidget):
    """
    Управление списком типов этикеток с возможностью редактирования
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.list_widget = QListWidget()
        self.layout.addWidget(self.list_widget)

        btn_layout = QHBoxLayout()
        self.layout.addLayout(btn_layout)

        self.add_btn = QPushButton("Добавить")
        self.remove_btn = QPushButton("Удалить")
        btn_layout.addWidget(self.add_btn)
        btn_layout.addWidget(self.remove_btn)

        self.add_btn.clicked.connect(self.add_label_type)
        self.remove_btn.clicked.connect(self.remove_label_type)

        # Начальный список
        self.load_defaults()

    def load_defaults(self):
        defaults = ["КОРПУС", "ФАСАДЫ МДФ", "ФАСАДЫ ПЛАСТИК", "Профиль/доп элемент", "Оргалит"]
        self.list_widget.clear()
        for d in defaults:
            item = QListWidgetItem(d)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.list_widget.addItem(item)

    def add_label_type(self):
        item = QListWidgetItem("Новый тип")
        item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
        self.list_widget.addItem(item)
        self.list_widget.editItem(item)

    def remove_label_type(self):
        selected = self.list_widget.selectedItems()
        if not selected:
            return
        for item in selected:
            self.list_widget.takeItem(self.list_widget.row(item))

    def get_label_types(self):
        types = []
        for i in range(self.list_widget.count()):
            text = self.list_widget.item(i).text().strip()
            if text:
                types.append(text)
        return types


class LabelOrderWidget(QWidget):
    """
    Виджет для выбора типа этикетки и количества
    """
    def __init__(self, label_types, parent=None):
        super().__init__(parent)
        self.layout = QHBoxLayout()
        self.setLayout(self.layout)

        self.combo = QComboBox()
        self.combo.addItems(label_types)
        self.spin = QSpinBox()
        self.spin.setMinimum(1)
        self.spin.setMaximum(1000)
        self.spin.setValue(1)

        self.remove_btn = QPushButton("Удалить")

        self.layout.addWidget(QLabel("Тип этикетки:"))
        self.layout.addWidget(self.combo)
        self.layout.addWidget(QLabel("Количество:"))
        self.layout.addWidget(self.spin)
        self.layout.addWidget(self.remove_btn)

        self.remove_btn.clicked.connect(self.delete_self)

    def delete_self(self):
        self.setParent(None)
        self.deleteLater()

    def get_data(self):
        return self.combo.currentText(), self.spin.value()

    def set_label_types(self, label_types):
        current = self.combo.currentText()
        self.combo.clear()
        self.combo.addItems(label_types)
        if current in label_types:
            self.combo.setCurrentText(current)


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Генератор этикеток")

        self.settings_path = "settings.json"
        self.settings = self.load_settings()

        self.order_info = None

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # --- Файл раскроя ---
        file_layout = QHBoxLayout()
        self.layout.addLayout(file_layout)

        self.file_edit = QLineEdit()
        self.file_edit.setPlaceholderText("Путь к файлу РАСКРОЙ 2025.xlsx")
        file_layout.addWidget(self.file_edit)

        self.browse_btn = QPushButton("Обзор")
        file_layout.addWidget(self.browse_btn)

        self.remember_checkbox = QCheckBox("Запомнить путь")
        file_layout.addWidget(self.remember_checkbox)

        self.browse_btn.clicked.connect(self.browse_file)

        # Если в настройках есть сохранённый путь, подставляем
        if self.settings.get('file_path'):
            self.file_edit.setText(self.settings['file_path'])
            self.remember_checkbox.setChecked(True)

        # --- Номер заказа ---
        order_layout = QHBoxLayout()
        self.layout.addLayout(order_layout)

        order_layout.addWidget(QLabel("Номер заказа:"))
        self.order_edit = QLineEdit()
        order_layout.addWidget(self.order_edit)

        self.search_btn = QPushButton("Найти")
        order_layout.addWidget(self.search_btn)
        self.search_btn.clicked.connect(self.search_order)

        # --- Информация о заказе (редактируемая) ---
        self.info_group = QGroupBox("Информация о заказе")
        self.info_layout = QFormLayout()
        self.info_group.setLayout(self.info_layout)
        self.layout.addWidget(self.info_group)

        self.fields = {}
        for field in ['store_application_number', 'client', 'full_name', 'item_name', 'width', 'height', 'depth', 'carcase', 'extra_component', 'facade', 'weight']:
            le = QLineEdit()
            self.fields[field] = le
            label = {
                'store_application_number': "№ магазина / заявка",
                'client': "Клиент",
                'full_name': "Полное наименование",
                'item_name': "Наименование изделия",
                'width': "Ширина (мм)",
                'height': "Высота (мм)",
                'depth': "Глубина (мм)",
                'carcase': "Корпус",
                'extra_component': "Доп. компонент",
                'facade': "Фасад",
                'weight': "Вес (кг)"
            }[field]
            self.info_layout.addRow(label + ":", le)

        # --- Управление типами этикеток ---
        self.label_type_manager = LabelTypeManager()
        self.layout.addWidget(self.label_type_manager)

        # --- Добавление этикеток ---
        self.labels_layout = QVBoxLayout()
        self.layout.addLayout(self.labels_layout)

        self.add_label_btn = QPushButton("Добавить тип этикеток")
        self.layout.addWidget(self.add_label_btn)
        self.add_label_btn.clicked.connect(self.add_label_order)

        # --- Кнопка создания ---
        create_layout = QHBoxLayout()
        self.layout.addLayout(create_layout)

        self.create_btn = QPushButton("Создать этикетки")
        create_layout.addWidget(self.create_btn)
        self.create_btn.clicked.connect(self.create_labels)

        # --- Список виджетов для заказов этикеток ---
        self.label_order_widgets = []

        # Если запомнили путь, загрузим
        self.file_path = self.file_edit.text()

    def browse_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "Выберите файл РАСКРОЙ 2025", "", "Excel Files (*.xlsx *.xls)")
        if path:
            self.file_edit.setText(path)
            if self.remember_checkbox.isChecked():
                self.save_settings({'file_path': path})

    def load_settings(self):
        if os.path.exists(self.settings_path):
            try:
                with open(self.settings_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except:
                return {}
        return {}

    def save_settings(self, data):
        self.settings.update(data)
        with open(self.settings_path, "w", encoding="utf-8") as f:
            json.dump(self.settings, f, ensure_ascii=False, indent=4)

    def search_order(self):
        file_path = self.file_edit.text().strip()
        if not file_path:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите файл РАСКРОЙ 2025.")
            return
        order_number = self.order_edit.text().strip()
        if not order_number:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, введите номер заказа.")
            return

        try:
            processor = OrderProcessor(file_path)
            order_info = processor.process_order(order_number)
            self.order_info = order_info

            # Заполняем поля
            self.fields['store_application_number'].setText(str(order_info.store_application_number))
            self.fields['client'].setText(str(order_info.client))
            self.fields['full_name'].setText(str(order_info.full_name))
            self.fields['item_name'].setText(str(order_info.item_name))
            self.fields['width'].setText(str(order_info.dimensions[0]))
            self.fields['height'].setText(str(order_info.dimensions[1]))
            self.fields['depth'].setText(str(order_info.dimensions[2]))
            self.fields['carcase'].setText(str(order_info.carcase))
            self.fields['extra_component'].setText(str(order_info.extra_component or ''))
            self.fields['facade'].setText(str(order_info.facade or ''))
            self.fields['weight'].setText(str(int(order_info.weight)) if order_info.weight else '')

            QMessageBox.information(self, "Успех", "Заказ найден и загружен.")
        except FileNotFoundError:
            QMessageBox.critical(self, "Ошибка", "Файл не найден.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def add_label_order(self):
        label_types = self.label_type_manager.get_label_types()
        if not label_types:
            QMessageBox.warning(self, "Ошибка", "Список типов этикеток пуст.")
            return
        widget = LabelOrderWidget(label_types)
        self.labels_layout.addWidget(widget)
        self.label_order_widgets.append(widget)

    def create_labels(self):
        if not self.order_info:
            QMessageBox.warning(self, "Ошибка", "Сначала найдите заказ.")
            return
        if not self.label_order_widgets:
            QMessageBox.warning(self, "Ошибка", "Добавьте хотя бы один тип этикеток.")
            return

        # Собираем данные для этикеток
        label_orders = []
        for widget in self.label_order_widgets:
            label_type, count = widget.get_data()
            if count <= 0:
                QMessageBox.warning(self, "Ошибка", "Количество этикеток должно быть положительным.")
                return
            label_orders.append((label_type, count))

        # Запрос пути сохранения
        save_path, _ = QFileDialog.getSaveFileName(self, "Сохранить файл с этикетками", "multylabel.xlsx", "Excel Files (*.xlsx)")
        if not save_path:
            return

        # Создаем книгу
        wb = Workbook()
        ws = wb.active

        # Установка ширины столбцов один раз
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        current_row = 1
        total_labels = sum(count for _, count in label_orders)
        label_index = 1

        # Собираем редактируемые значения из полей
        editable_values = {
            'store_application_number': self.fields['store_application_number'].text(),
            'client': self.fields['client'].text(),
            'full_name': self.fields['full_name'].text(),
            'item_name': self.fields['item_name'].text(),
            'width': self.fields['width'].text(),
            'height': self.fields['height'].text(),
            'depth': self.fields['depth'].text(),
            'carcase': self.fields['carcase'].text(),
            'extra_component': self.fields['extra_component'].text(),
            'facade': self.fields['facade'].text(),
            'weight': self.fields['weight'].text(),
        }

        # Попытка преобразовать размеры и вес в числа, если возможно
        try:
            editable_values['width'] = int(editable_values['width'])
        except:
            editable_values['width'] = 0
        try:
            editable_values['height'] = int(editable_values['height'])
        except:
            editable_values['height'] = 0
        try:
            editable_values['depth'] = int(editable_values['depth'])
        except:
            editable_values['depth'] = 0
        try:
            editable_values['weight'] = float(editable_values['weight'])
        except:
            editable_values['weight'] = None

        # Создаем этикетки
        for label_type, count in label_orders:
            for i in range(1, count + 1):
                create_label(ws, current_row, self.order_info, label_type, total_labels, label_index, editable_values)
                current_row += 17
                label_index += 1

        try:
            wb.save(save_path)
            QMessageBox.information(self, "Готово", f"Файл с этикетками сохранён: {save_path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл: {e}")

def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.resize(900, 700)
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
